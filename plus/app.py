"""
내븐 Corporation — 내부 DB Heartbeat 진단 시스템 (CVE-2014-0160 시뮬레이션)

기본 동작:
  client → {payload, length} → server
  server → payload[:length] 만 반환 (정상)

취약 동작 (Heartbleed):
  length > len(payload) 이면
  server → payload + MEMORY[: length - len(payload)] 반환

MEMORY = [가짜 힙 가비지][개인정보.xlsx 원본 바이트][가짜 로그]

leak_size >= HOOK_THRESHOLD (= 가비지 + xlsx 크기) 이면
  학생이 누출된 바이트만으로 xlsx 전체를 복원할 수 있는 상태 → /api/leaked-file 로
  원본 개인정보.xlsx 를 그대로 내려줌 (hook).
"""
from flask import Flask, request, jsonify, send_from_directory, send_file
from pathlib import Path

BASE = Path(__file__).parent
DATA_DIR = BASE / 'data'
XLSX_PATH = DATA_DIR / '개인정보.xlsx'
IMAGES_DIR = BASE / 'images'  # plus/images/NAVEN.png

app = Flask(__name__, static_folder=str(BASE))

MAX_LENGTH = 65535  # 실제 Heartbleed 의 length 필드 한계 (uint16)


# -------------------------
# 가짜 개인정보.xlsx 생성
# -------------------------
def init_xlsx():
    DATA_DIR.mkdir(exist_ok=True)
    if XLSX_PATH.exists():
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = '내븐_고객DB'

    headers = ['번호', '이름', '주민등록번호', '휴대전화', '이메일', '주소', '월 결제액(₩)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='B91C1C')
        cell.alignment = Alignment(horizontal='center')

    rows = [
        ('김민서', '880101-2******', '010-2341-7782', 'kim.ms@navenmail.com', '서울 강남구 테헤란로 31', 47800),
        ('박정훈', '910415-1******', '010-4521-3398', 'park.jh@navenmail.com', '서울 마포구 와우산로 94', 88200),
        ('이수아', '950822-2******', '010-8847-2210', 'lee.sa@navenmail.com', '경기 성남시 분당구 정자일로 8', 32500),
        ('최도윤', '870307-1******', '010-3344-9921', 'choi.dy@navenmail.com', '인천 연수구 송도과학로 12', 119900),
        ('정하린', '930521-2******', '010-7712-4480', 'jung.hr@navenmail.com', '부산 해운대구 센텀중앙로 79', 54200),
        ('한지우', '900915-1******', '010-4421-8830', 'han.jw@navenmail.com', '대전 유성구 대학로 99', 76100),
        ('윤서연', '961203-2******', '010-9921-3318', 'yoon.sy@navenmail.com', '서울 종로구 대학로 57', 41200),
        ('장민혁', '850628-1******', '010-5582-1129', 'jang.mh@navenmail.com', '경기 수원시 영통구 광교로 42', 152800),
        ('조예린', '981114-2******', '010-3308-9974', 'cho.yr@navenmail.com', '서울 송파구 올림픽로 240', 28900),
        ('서지호', '920430-1******', '010-2218-6650', 'seo.jh@navenmail.com', '광주 북구 첨단과기로 123', 64500),
        ('강민채', '940707-2******', '010-7770-4421', 'kang.mc@navenmail.com', '대구 수성구 동대구로 366', 83400),
        ('오태형', '861119-1******', '010-1129-5582', 'oh.th@navenmail.com', '서울 영등포구 여의대로 24', 211400),
        ('임수빈', '991008-2******', '010-4480-7712', 'lim.sb@navenmail.com', '경기 고양시 일산동구 정발산로 22', 33700),
        ('홍재민', '890225-1******', '010-6650-2218', 'hong.jm@navenmail.com', '서울 서초구 강남대로 373', 98600),
        ('문아윤', '970716-2******', '010-9974-3308', 'moon.ay@navenmail.com', '경남 창원시 의창구 중앙대로 245', 47100),
        ('신유찬', '910303-1******', '010-1129-7782', 'shin.yc@navenmail.com', '서울 노원구 동일로 1413', 61300),
        ('황지안', '950820-2******', '010-2341-9921', 'hwang.ja@navenmail.com', '경기 성남시 수정구 산성대로 451', 52800),
        ('백건우', '880411-1******', '010-4421-2210', 'baek.gw@navenmail.com', '서울 마포구 양화로 45', 142200),
        ('남시현', '930129-2******', '010-7712-9921', 'nam.sh@navenmail.com', '인천 부평구 부평대로 283', 39400),
        ('구지훈', '961231-1******', '010-3344-7782', 'gu.jh@navenmail.com', '울산 남구 삼산로 233', 73100),
        ('류세아', '990418-2******', '010-5582-3318', 'ryu.sa@navenmail.com', '서울 강서구 공항대로 376', 29800),
        ('전민준', '870924-1******', '010-1129-4480', 'jeon.mj@navenmail.com', '대전 서구 둔산대로 117', 156700),
        ('고채원', '940215-2******', '010-7770-1129', 'go.cw@navenmail.com', '경기 안양시 동안구 시민대로 235', 44900),
        ('권태이', '910811-1******', '010-2218-9974', 'kwon.ti@navenmail.com', '서울 광진구 자양로 117', 67200),
        ('나하은', '981105-2******', '010-9921-5582', 'na.he@navenmail.com', '부산 사하구 낙동남로 1413', 35600),
        ('도현우', '860620-1******', '010-4480-2218', 'do.hw@navenmail.com', '서울 용산구 한강대로 405', 188300),
        ('마유나', '950312-2******', '010-6650-4421', 'ma.yn@navenmail.com', '경기 용인시 기흥구 강남로 9', 41700),
        ('배성호', '890717-1******', '010-9974-7712', 'bae.sh@navenmail.com', '대구 달서구 달구벌대로 1095', 84500),
        ('사혜진', '970208-2******', '010-3308-1129', 'sa.hj@navenmail.com', '서울 동작구 노량진로 84', 38200),
        ('안재현', '920525-1******', '010-1129-3344', 'ahn.jh@navenmail.com', '인천 남동구 인하로 425', 71900),
        ('어수현', '960927-2******', '010-2341-5582', 'eo.sh@navenmail.com', '경기 부천시 원미구 길주로 210', 49300),
        ('지민결', '880803-1******', '010-4421-7770', 'ji.mg@navenmail.com', '서울 성북구 정릉로 77', 132400),
        ('차예진', '940110-2******', '010-7712-2218', 'cha.yj@navenmail.com', '경기 화성시 동탄대로 537', 53600),
        ('탁승원', '910612-1******', '010-3344-9921', 'tak.sw@navenmail.com', '서울 동대문구 회기로 26', 79800),
        ('편소율', '990229-2******', '010-5582-4480', 'pyeon.sr@navenmail.com', '강원 춘천시 공지로 126', 31100),
        ('하준우', '870505-1******', '010-1129-6650', 'ha.jw@navenmail.com', '서울 중구 명동길 73', 174500),
        ('갈민서', '951018-2******', '010-7770-3308', 'gal.ms@navenmail.com', '경기 의정부시 평화로 549', 46200),
        ('남궁윤호', '890901-1******', '010-2218-1129', 'namgung.yh@navenmail.com', '서울 은평구 통일로 684', 92700),
        ('독고연아', '961107-2******', '010-9921-7712', 'dokgo.ya@navenmail.com', '제주 제주시 첨단로 213', 37800),
        ('황보재경', '930724-1******', '010-4480-3344', 'hwangbo.jk@navenmail.com', '경기 김포시 김포한강9로 76', 108300),
    ]

    for i, r in enumerate(rows, start=1):
        ws.append((i, *r))

    widths = [6, 10, 18, 16, 28, 36, 14]
    for col, w in zip('ABCDEFG', widths):
        ws.column_dimensions[col].width = w

    wb.save(XLSX_PATH)


init_xlsx()


# -------------------------
# 시뮬레이션 메모리 구성
# -------------------------
def build_memory():
    """[가짜 힙][개인정보.xlsx 바이트][가짜 로그] — 학생이 hex dump 에서
    PK 시그니처 + `file_size=<숫자>` ASCII 단서를 발견할 수 있도록.
    """
    xlsx_bytes = XLSX_PATH.read_bytes()

    # 학생이 hex dump 오른쪽 ASCII 컬럼에서 발견해야 할 단서.
    # 16-bit zero-pad 16진수로 박아둠 — 학생은 이걸 10진수로 환산해서 length 로 넣어야 함.
    size_marker = f'file_size={format(len(xlsx_bytes), "04x")}\x00'.encode()

    garbage_before = (
        b'NAVEN_HEAP_SEG[0x7fff8000]\x00\x00'
        b'session_admin=eyJ1IjoiYWRtaW4iLCJyIjoiYWRtaW4ifQ\x00'
        b'last_query=SELECT * FROM customers WHERE id>0\x00'
        + size_marker
        + (b'\xde\xad\xbe\xef' * 8)  # 명시적 + : '*' 가 앞 리터럴들까지 묶지 않도록
    )
    garbage_before = garbage_before.ljust(200, b'\x00')

    garbage_after = (
        b'\n[18:23:11] heartbeat OK len=2'
        b'\n[18:23:14] /api/admin/document accessed by user_id=2'
        b'\n[18:25:02] memory page swapped'
    ).ljust(1024, b'\x00')

    return garbage_before, xlsx_bytes, garbage_after


GARBAGE_BEFORE, XLSX_BYTES, GARBAGE_AFTER = build_memory()
MEMORY = GARBAGE_BEFORE + XLSX_BYTES + GARBAGE_AFTER
XLSX_SIZE = len(XLSX_BYTES)  # hook 정답 길이 (가비지에 박힌 file_size= 값과 동일)


# -------------------------
# 정적 파일 서빙
# -------------------------
@app.route('/')
def index():
    return send_from_directory(str(BASE), 'index.html')


@app.route('/<path:path>')
def static_files(path):
    return send_from_directory(str(BASE), path)


@app.route('/images/<path:filename>')
def images(filename):
    return send_from_directory(str(IMAGES_DIR), filename)


# -------------------------
# Heartbeat (의도적 취약)
# -------------------------
@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    data = request.get_json(silent=True) or {}
    payload = str(data.get('payload', ''))

    try:
        length = int(data.get('length', 0))
    except (TypeError, ValueError):
        return jsonify({'error': '응답 길이는 정수여야 합니다.'}), 400

    if length < 0 or length > MAX_LENGTH:
        return jsonify({'error': f'응답 길이는 0~{MAX_LENGTH} 범위여야 합니다.'}), 400

    payload_bytes = payload.encode('utf-8')

    # 정상: payload 만 잘라서 에코
    response = payload_bytes[:length]
    leak_size = 0

    # 취약: length > payload → 메모리에서 끌어옴
    if length > len(payload_bytes):
        leak_size = length - len(payload_bytes)
        if leak_size <= len(MEMORY):
            leaked = MEMORY[:leak_size]
        else:
            leaked = MEMORY + b'\x00' * (leak_size - len(MEMORY))
        response = payload_bytes + leaked

    # 학생이 누출된 ASCII 단서(`file_size=...`)를 발견하고
    # 그 값을 정확히 length 로 넣을 때만 hook 발동.
    hook = (length == XLSX_SIZE)

    return jsonify({
        'payload_size': len(payload_bytes),
        'requested_length': length,
        'response_size': len(response),
        'leak_size': leak_size,
        'response_hex': response.hex(),
        'hook': hook,
    })


# -------------------------
# Hook: 누출 충분 시 원본 xlsx 전달
# -------------------------
@app.route('/api/leaked-file')
def leaked_file():
    return send_file(
        str(XLSX_PATH),
        as_attachment=True,
        download_name='개인정보.xlsx',
    )


if __name__ == '__main__':
    print(f'[init] xlsx 크기: {XLSX_SIZE} bytes')
    print(f'[init] MEMORY 크기: {len(MEMORY)} bytes')
    print(f'[init] hook 정답 length: {XLSX_SIZE} (가비지에 박힌 file_size= 값)')
    print('서버 시작: http://localhost:5001')
    app.run(debug=True, port=5001)
